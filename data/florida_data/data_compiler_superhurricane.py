#!/usr/bin/env python 

import sys
import os
from openpyxl import load_workbook
import csv

WORKBOOK_FILENAME = './Florida-hurricane-data.xlsx'
ALIAS = {
    'New West Palm Beach Distribution CC':          'PalmBeach',
    'Jacksonville Tierpoint':                       'Jacksonville',
    'Miami Equinix MI1':                            'MiamiMI1',
    'Miami Dade County Resource Recovery Facility': 'DadeCounty',
    'Riviera':                                      'Riviera',
    'Cape Canaveral':                               'CapeCanaveral',
    'Port Orange':                                  'PortOrange',
    'Tampa Cyxtera':                                'TampaCyxtera',
    'Sarasota/Tampa':                               'SarasotaTampa',
    'Boca Raton Equinix MI3':                       'BocaRatonMI3'
}
DC_LIST = ['Jacksonville Tierpoint', 'Miami Equinix MI1', 'Tampa Cyxtera', 'Boca Raton Equinix MI3']
CC_LIST = list(set(ALIAS.keys()) - set(DC_LIST))

# CC1_LIST = CC_LIST
# CC2_LIST = CC_LIST
# DC1_LIST = DC_LIST
# DC2_LIST = DC_LIST
# DC3_LIST = DC_LIST
CC1_LIST = ['New West Palm Beach Distribution CC']
CC2_LIST = ['Port Orange']
DC1_LIST = ['Jacksonville Tierpoint', 'Miami Equinix MI1']
DC2_LIST = DC1_LIST
DC3_LIST = DC1_LIST

HURR_RES_MIN_COL = 8
HURR_RES_MAX_COL = 1007

def main():
    # Delete existing csv files
    directory = "./"
    files_in_directory = os.listdir(directory)
    filtered_files = [file for file in files_in_directory if file.endswith(".csv")]
    for file in filtered_files:
        path_to_file = os.path.join(directory, file)
        os.remove(path_to_file)

    # Delete existing stats files
    directory = "./"
    files_in_directory = os.listdir(directory)
    filtered_files = [file for file in files_in_directory if file.endswith(".stats")]
    for file in filtered_files:
        path_to_file = os.path.join(directory, file)
        os.remove(path_to_file)


    workbook = load_workbook(filename=WORKBOOK_FILENAME)

    hurricane_names = ['FL_superhurricane']
    combinations_done = []
    for hid in range(len(hurricane_names)):
        cur_hname = hurricane_names[hid]
        workbook.active = 2 # third sheet in order
        sheet = workbook.active

        col_A = sheet["A"]

        # print(col_A)

        for CC1 in CC1_LIST:
            for CC2 in CC2_LIST:
                for DC1 in DC1_LIST:
                    for DC2 in DC2_LIST:
                        for DC3 in DC3_LIST:
                            this_combination = ((CC1, CC2), (DC1, DC2, DC3))
                            if no_need(this_combination) == False:
                                if not already_done(this_combination, combinations_done):
                                    #print(CC1, CC2, DC1, DC2, DC3)

                                    for cell_id in range(len(col_A)):
                                        if col_A[cell_id].value == CC1:
                                            CC1_row = cell_id + 1
                                        if col_A[cell_id].value == CC2:
                                            CC2_row = cell_id + 1
                                        if col_A[cell_id].value == DC1:
                                            DC1_row = cell_id + 1
                                        if col_A[cell_id].value == DC2:
                                            DC2_row = cell_id + 1
                                        if col_A[cell_id].value == DC3:
                                            DC3_row = cell_id + 1

                                    #print(CC1_row, CC2_row, DC1_row, DC2_row, DC3_row)

                                    ids = ["ID"] + list(range(1,1001))

                                    for row in sheet.iter_rows(min_row=CC1_row, max_row=CC1_row, min_col=HURR_RES_MIN_COL, max_col=HURR_RES_MAX_COL, values_only=True):
                                        CC1_row_values = ["CC1"] + list(row)

                                    for row in sheet.iter_rows(min_row=CC2_row, max_row=CC2_row, min_col=HURR_RES_MIN_COL, max_col=HURR_RES_MAX_COL, values_only=True):
                                        CC2_row_values = ["CC2"] + list(row)

                                    for row in sheet.iter_rows(min_row=DC1_row, max_row=DC1_row, min_col=HURR_RES_MIN_COL, max_col=HURR_RES_MAX_COL, values_only=True):
                                        DC1_row_values = ["DC1"] + list(row)

                                    for row in sheet.iter_rows(min_row=DC2_row, max_row=DC2_row, min_col=HURR_RES_MIN_COL, max_col=HURR_RES_MAX_COL, values_only=True):
                                        DC2_row_values = ["DC2"] + list(row)

                                    for row in sheet.iter_rows(min_row=DC3_row, max_row=DC3_row, min_col=HURR_RES_MIN_COL, max_col=HURR_RES_MAX_COL, values_only=True):
                                        DC3_row_values = ["DC3"] + list(row)

                                    csv_filename = cur_hname + "_" + ALIAS[CC1] + "_" + ALIAS[CC2] + "_" + ALIAS[DC1] + "_" + ALIAS[DC2] + "_" + ALIAS[DC3] + ".csv"

                                    print(csv_filename)

                                    with open(csv_filename, 'w', newline='') as csv_file:
                                        writer = csv.writer(csv_file)
                                        writer.writerow(ids)
                                        writer.writerow(CC1_row_values)
                                        writer.writerow(CC2_row_values)
                                        writer.writerow(DC1_row_values)
                                        writer.writerow(DC2_row_values)
                                        writer.writerow(DC3_row_values)


                                    with open(csv_filename+".stats", 'w') as f: 
                                        i = 0
                                        for row_vals in [CC1_row_values, CC2_row_values, DC1_row_values, DC2_row_values, DC3_row_values]:
                                            nums = [float(n) for n in row_vals[1:]]
                                        
                                            percent = str(round(((sum(nums)/len(nums))*100), 2)) + "%"
                                            if i == 0:
                                                key = 'CC1'
                                            elif i == 1:
                                                key = 'CC2'
                                            elif i == 2:
                                                key = 'DC1'
                                            elif i == 3:
                                                key = 'DC2'
                                            elif i == 4:
                                                key = 'DC3'
                                            i += 1
                                            f.write('%s: %s\n' % (key, percent))
                                        
                                        # 'CC1andCC2':
                                        CC1andCC2 = [bool(a) and bool(b) for a, b in zip(CC1_row_values[1:], CC2_row_values[1:])]
                                        CC1andCC2 = [int(x) for x in CC1andCC2]
                                        
                                        # 'CC1orCC2':
                                        CC1orCC2 = [bool(a) or bool(b) for a, b in zip(CC1_row_values[1:], CC2_row_values[1:])]
                                        CC1orCC2 = [int(x) for x in CC1orCC2]

                                        # 'CC1andCC2andDC1':
                                        CC1andCC2andDC1 = [bool(a) and bool(b) for a, b in zip(CC1andCC2, DC1_row_values[1:])]
                                        CC1andCC2andDC1 = [int(x) for x in CC1andCC2andDC1]

                                        # 'CC1orDC1':
                                        CC1orDC1 = [bool(a) or bool(b) for a, b in zip(CC1_row_values[1:], DC1_row_values[1:])]
                                        CC1orDC1 = [int(x) for x in CC1orDC1]
                                        
                                        # 'CC2orDC1':
                                        CC2orDC1 = [bool(a) or bool(b) for a, b in zip(CC2_row_values[1:], DC1_row_values[1:])]
                                        CC2orDC1 = [int(x) for x in CC2orDC1]
                                        
                                        f.write('%s: %s\n' % ('CC1andCC2', str(round(((sum(CC1andCC2)/len(CC1andCC2))*100), 2)) + "%"))
                                        f.write('%s: %s\n' % ('CC1orCC2', str(round(((sum(CC1orCC2)/len(CC1orCC2))*100), 2)) + "%"))
                                        f.write('%s: %s\n' % ('CC1andCC2andDC1', str(round(((sum(CC1andCC2andDC1)/len(CC1andCC2andDC1))*100), 2)) + "%"))
                                        f.write('%s: %s\n' % ('CC1orDC1', str(round(((sum(CC1orDC1)/len(CC1orDC1))*100), 2)) + "%"))
                                        f.write('%s: %s\n' % ('CC2orDC1', str(round(((sum(CC2orDC1)/len(CC2orDC1))*100), 2)) + "%"))

def no_need(this_combination):
    # CCs = this_combination[0]
    # DCs = this_combination[1]

    # no_need = False

    # if CCs[0] == CCs[1]:
    #     no_need = True
    # if DCs[0] == DCs[1]:
    #     no_need = True
    # if DCs[0] == DCs[2]:
    #     no_need = True
    # if DCs[1] == DCs[2]:
    #     no_need = True
    
    # return no_need
    return False

def already_done(this_combination, combinations_done):
    CCs = this_combination[0]
    DCs = this_combination[1]

    CC_order_1 = (CCs[0], CCs[1])
    CC_order_2 = (CCs[1], CCs[0])
    
    DC_order_1 = (DCs[0], DCs[1], DCs[2])
    DC_order_2 = (DCs[0], DCs[2], DCs[1])
    DC_order_3 = (DCs[1], DCs[0], DCs[2])
    DC_order_4 = (DCs[1], DCs[2], DCs[0])
    DC_order_5 = (DCs[2], DCs[0], DCs[1])
    DC_order_6 = (DCs[2], DCs[1], DCs[0])

    already_done = False
    for DC_comb in [DC_order_1, DC_order_2, DC_order_3, DC_order_4, DC_order_5, DC_order_6]:
        this_comb_op1 = (CC_order_1, DC_comb)
        this_comb_op2 = (CC_order_2, DC_comb)
        if (this_comb_op1 in combinations_done) or (this_comb_op2 in combinations_done):
            already_done = True

    return already_done

if __name__ == '__main__':
    main()