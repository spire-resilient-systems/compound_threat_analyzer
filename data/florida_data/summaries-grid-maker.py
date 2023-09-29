#!/usr/bin/env python 

from openpyxl import load_workbook
import csv

WORKBOOK_FILENAME = 'Florida-hurricane-data.xlsx'
ALIAS = {
    'New West Palm Beach Distribution CC':          'PalmBeach',
    'Jacksonville Tierpoint':                       'Jacksonville',
    'Miami Equinix MI1':                            'MiamiMI1',
    'Miami Dade County Resource Recovery Facility': 'DadeCounty',
    'Riviera':                                      'Riviera',
    'Cape Canaveral':                               'CapeCanaveral',
    'Port Orange':                                  'PortOrange',
    'Tampa Cyxtera':                                'TampaCyxtera',
    'Sarasota/Tampa':                               'SarasotaTampa',
    'Boca Raton Equinix MI3':                       'BocaRatonMI3'
}

SITES = ALIAS.keys()

HURR_RES_MIN_COL = 8
HURR_RES_MAX_COL = 1007

workbook = load_workbook(filename=WORKBOOK_FILENAME)

with open('summaries_ands.csv', 'w') as f: 
    f.write('') # empty it out
with open('summaries_ors.csv', 'w') as f: 
    f.write('') # empty it out

combinations = []
for site1 in SITES:
    for site2 in SITES:
        this_combination = (site1, site2)
        combinations.append(this_combination)

summaries_and = []
summaries_or = []
summaries_corr = []
for combination in combinations:
    print(combination)
    site1, site2 = combination[0], combination[1]

    workbook.active = 2 # third sheet in order
    sheet = workbook.active

    col_A = sheet["A"]

    for cell_id in range(len(col_A)):
        if col_A[cell_id].value == site1:
            site1_row = cell_id + 1
        if col_A[cell_id].value == site2:
            site2_row = cell_id + 1

    site1_row_values = list(list(sheet.iter_rows(min_row=site1_row, max_row=site1_row, min_col=HURR_RES_MIN_COL, max_col=HURR_RES_MAX_COL, values_only=True))[0])
    site2_row_values = list(list(sheet.iter_rows(min_row=site2_row, max_row=site2_row, min_col=HURR_RES_MIN_COL, max_col=HURR_RES_MAX_COL, values_only=True))[0])

    # site1 AND site2:
    ands = [bool(a) and bool(b) for a, b in zip(site1_row_values, site2_row_values)]
    ands = [int(x) for x in ands]
    
    # site1 OR  site2:
    ors = [bool(a) or bool(b) for a, b in zip(site1_row_values, site2_row_values)]
    ors = [int(x) for x in ors]    

    # correlations site 1 with site 2:
    # 1 means site failure. 0 means site is running fine.
    count_correlated_failure = 0
    index = 0
    for v in site1_row_values:
        # in the excel file, 1 means site is down
        if int(v) == 1: # if row site is down
            if site2_row_values[index] == 1:
                count_correlated_failure += 1
        index += 1
    site2_total_fails = 0
    for v in site2_row_values:
        site2_total_fails += v

    and_perc = str(round(((sum(ands)/len(ands))*100), 2)) + "%"
    or_perc = str(round(((sum(ors)/len(ors))*100), 2)) + "%"
    try:
        count_correlated_failure_perc = count_correlated_failure/site2_total_fails
    except ZeroDivisionError:
        count_correlated_failure_perc = 'DivByZero'

    summaries_and.append((ALIAS[site1], ALIAS[site2], and_perc))
    summaries_or.append((ALIAS[site1], ALIAS[site2], or_perc))
    summaries_corr.append((ALIAS[site1], ALIAS[site2], count_correlated_failure_perc))

def make_csv(filename, summary_obj):
    with open(filename, 'w+', newline ='') as f: 
        top_left_corner_cell = 'sites'
        # vals = (row_site, col_site, %_val)
        row_names = []
        col_names = []
        
        for val in summary_obj:
            if val[0] not in row_names:
                row_names.append(val[0])
            if val[1] not in col_names:
                col_names.append(val[1])
            
        csv_rows = []
        csv_row_len = len(col_names) + 1
        # top row:
        csv_rows.append([top_left_corner_cell])
        for c in col_names:
            csv_rows[0].append(c)
        # rest of the rows:
        for i in range(csv_row_len - 1):
            this_row = []
            this_row_name = row_names[i]
            this_row.append(this_row_name)

            for c in col_names:
                this_row_col_val = None
                for v in summary_obj:
                    if v[0] == this_row_name and v[1] == c:
                        this_row_col_val = v[2]
                        break
                this_row.append(this_row_col_val)

            csv_rows.append(this_row)

        # for r in csv_rows:
        #     print(r)
        # f.write("\n".join(",".join(csv_rows)))
        write = csv.writer(f)
        write.writerows(csv_rows)

make_csv('summaries_ands.csv', summaries_and)
make_csv('summaries_ors.csv', summaries_or)
make_csv('summaries_corrs.csv', summaries_corr)