#!/usr/bin/env python

# Data Compiler

import sys
import os
from openpyxl import load_workbook
import csv

# Delete existing csv files
directory = "./"
files_in_directory = os.listdir(directory)
filtered_files = [file for file in files_in_directory if file.endswith(".csv")]
for file in filtered_files:
    path_to_file = os.path.join(directory, file)
    os.remove(path_to_file)


workbook = load_workbook(filename="Hawaii-hurricane-data.xlsx")
#print(workbook.sheetnames)

hurricane_names = workbook.sheetnames

CC1_list = ["honolulu"]
CC2_list = ["waiau", "kahe", "aes-hawaii", "kalaeloa"]
DC1_list = ["DRFotress"]
DC2_list = ["AlohaNap"]
DC3_list = ["MTP"]

for hid in range(len(hurricane_names)):
    cur_hname = hurricane_names[hid]
    workbook.active = hid

    sheet = workbook.active

    if cur_hname != sheet.title:
        sys.exit("Error: Current sheet name does not match workbook active title!")

    sheet_C = sheet["C"]

    for CC1 in CC1_list:
        for CC2 in CC2_list:
            for DC1 in DC1_list:
                for DC2 in DC2_list:
                    for DC3 in DC3_list:

                        #print(CC1, CC2, DC1, DC2, DC3)

                        for cell_id in range(len(sheet_C)):
                            if sheet_C[cell_id].value == CC1:
                                CC1_row = cell_id + 1
                            if sheet_C[cell_id].value == CC2:
                                CC2_row = cell_id + 1
                            if sheet_C[cell_id].value == DC1:
                                DC1_row = cell_id + 1
                            if sheet_C[cell_id].value == DC2:
                                DC2_row = cell_id + 1
                            if sheet_C[cell_id].value == DC3:
                                DC3_row = cell_id + 1

                        #print(CC1_row, CC2_row, DC1_row, DC2_row, DC3_row)

                        ids = ["ID"] + list(range(1,1001))

                        for row in sheet.iter_rows(min_row=CC1_row, max_row=CC1_row, min_col=15, max_col=1014, values_only=True):
                            CC1_row_values = ["CC1"] + list(row)

                        for row in sheet.iter_rows(min_row=CC2_row, max_row=CC2_row, min_col=15, max_col=1014, values_only=True):
                            CC2_row_values = ["CC2"] + list(row)

                        for row in sheet.iter_rows(min_row=DC1_row, max_row=DC1_row, min_col=15, max_col=1014, values_only=True):
                            DC1_row_values = ["DC1"] + list(row)

                        for row in sheet.iter_rows(min_row=DC2_row, max_row=DC2_row, min_col=15, max_col=1014, values_only=True):
                            DC2_row_values = ["DC2"] + list(row)

                        for row in sheet.iter_rows(min_row=DC3_row, max_row=DC3_row, min_col=15, max_col=1014, values_only=True):
                            DC3_row_values = ["DC3"] + list(row)

                        csv_filename = cur_hname + "_" + CC1 + "_" + CC2 + "_" + DC1 + "_" + DC2 + "_" + DC3 + ".csv"

                        print(csv_filename)

                        with open(csv_filename, 'w', newline='') as csv_file:
                            writer = csv.writer(csv_file)
                            writer.writerow(ids)
                            writer.writerow(CC1_row_values)
                            writer.writerow(CC2_row_values)
                            writer.writerow(DC1_row_values)
                            writer.writerow(DC2_row_values)
                            writer.writerow(DC3_row_values)
                        
